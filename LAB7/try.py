# import pandas as pd

# file="nodes.xlsx"
# xl = pd.ExcelFile(file)
# print(xl.sheet_names)
# print(f.read())

from openpyxl import load_workbook

workbook = load_workbook('nodes.xlsx')
# first_sheet = workbook.get_sheet_names()[0]
# worksheet = workbook.get_sheet_by_name(first_sheet)
sheet=workbook.active
l=[]
x=0
# print(sheet.iter_rows())
for row in sheet.iter_rows():
	temp=[]
	for cell in row:
		temp.append(cell.value)
	temp.pop(0)
	l.append(temp)
l.pop(0)
print(l)
workbook = load_workbook('distance.xlsx')
sheet=workbook.active
d=[]
x=0
for row in sheet.iter_rows():
	temp=[]
	for cell in row:
		if(cell.value=="N"):
			temp.append(-1.0)
			continue
		temp.append(cell.value)
	temp.pop(0)
	d.append(temp)
d.pop(0)
print(d)
# check out the last row
# for cell in row:
# 	print (cell)


# Reading an excel file using Python 
# import xlrd 
  
# # Give the location of the file 
# loc = ("distance.xlsx") 
# f=open("d.txt","w")
  
# # To open Workbook 
# wb = xlrd.open_workbook(loc) 
# sheet = wb.sheet_by_index(0) 
  
# # For row 0 and column 0
# # f.write("[")
# for i in range(1,sheet.nrows): 
# 	# f.write("[")
# 	for j in range(1,sheet.ncols): 
# 		if sheet.cell_value(i,j)=="N":
# 			f.write("-1.0,")
# 			continue
# 		f.write("%.1f,"%(sheet.cell_value(i,j)))
# 	f.write("\n")
# 	# f.write("]\n")
# # f.write("]")

# f.close() 
# f=open("d.txt","r")
# t1=[]
# d=[]
# while 1:
# 	t2=f.readline()
# 	if t2=="":break
# 	t1=list(t2.split(","))
# 	t3=[]
# 	for i in range(len(t1)-1):
# 		x=float(t1[i])
# 		t3.append(x)
# 	d.append(t3)
# print(d[2])
# f.close()
