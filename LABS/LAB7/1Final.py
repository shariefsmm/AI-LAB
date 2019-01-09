import numpy as np
import heapq
from random import randint
import math
from openpyxl import load_workbook

class Environment:
	def __init__(self):
		self.read_nodes()
		self.read_distances()
		self.source=0
		self.goal=13
	def read_distances(self):
		workbook = load_workbook('distance.xlsx')
		sheet=workbook.active
		self.d=[]
		for row in sheet.iter_rows():
			temp=[]
			for cell in row:
				if(cell.value=="N"):
					temp.append(-1.0)
					continue
				temp.append(cell.value)
			temp.pop(0)
			self.d.append(temp)
		self.d.pop(0)
		# for i in self.d:
		# 	print(i)

	def read_nodes(self):
		workbook = load_workbook('nodes.xlsx')
		sheet=workbook.active
		self.l=[]
		for row in sheet.iter_rows():
			temp=[]
			for cell in row:
					temp.append(cell.value)
			temp.pop(0)
			self.l.append(temp)
		self.l.pop(0)
		self.n=len(self.l)
		# print(self.l)
		# print(self.n)

class Agent:
	def __init__(self):
		self.v=np.zeros(s.n,dtype='int');
		self.p=[ -1 for i in range(0,s.n)]
		self.dist=[-1 for i in range(0,s.n)]
		self.h=[]
		self.BFS()

	def Distance(self,x,y):
		(X,Y)=(s.l[s.goal][0],s.l[s.goal][1])
		return math.sqrt( (x-X)**2 +(y-Y)**2)

	def Print(self,g):
		print("Total cost taken is %s"%(g))
		t=[]
		x=s.goal
		while x!=-1:
			t.append(x)
			# print(x)
			x=self.p[x]
		for i in range(len(t)-1,-1,-1):
			print(t[i]+1,s.l[t[i]])

	def BFS(self):
		heapq.heapify(self.h)
		(x,y)=(s.l[s.source][0],s.l[s.source][1])
		heapq.heappush(self.h,(0+self.Distance(x,y),(s.source,-1,0)))
		while self.h:
			(A,(node,P,g))=heapq.heappop(self.h)
			if self.v[node]==1:continue
			if self.v[node]==0:self.v[node]=1
			self.p[node]=P
			if node==s.goal:
				self.Print(g)
				break
			for i in range(0,s.n):
				n=s.d[node][i]
				if n!=-1.0:
					(x,y)=(s.l[i][0],s.l[i][1])
					heapq.heappush(self.h,(g+n+self.Distance(x,y),(i,node,g+n)))


s=Environment()
a=Agent()


