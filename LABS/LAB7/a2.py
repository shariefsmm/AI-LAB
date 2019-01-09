import numpy as np
import heapq
from random import randint
import math
from openpyxl import load_workbook

class Environment:
	def __init__(self):
		self.read_nodes()
		self.read_distances()
		self.source=0
		self.goal=13
		self.Bus=50.0
		self.Cycle=25.0
		self.budget=5.4
		self.rate=10.0
	def read_distances(self):
		workbook = load_workbook('distance.xlsx')
		sheet=workbook.active
		self.d=[]
		for row in sheet.iter_rows():
			temp=[]
			for cell in row:
				if(cell.value=="N"):
					temp.append(-1.0)
					continue
				temp.append(cell.value)
			temp.pop(0)
			self.d.append(temp)
		self.d.pop(0)
		# for i in self.d:
		# 	print(i)

	def read_nodes(self):
		workbook = load_workbook('nodes.xlsx')
		sheet=workbook.active
		self.l=[]
		for row in sheet.iter_rows():
			temp=[]
			for cell in row:
					temp.append(cell.value)
			temp.pop(0)
			self.l.append(temp)
		self.l.pop(0)
		self.n=len(self.l)
		# print(self.l)
		# print(self.n)

class Agent:
	def __init__(self):
		self.v=np.zeros(s.n,dtype='int');
		self.p=[ -1 for i in range(0,s.n)]
		self.h=[]
		self.BFS()

	def Print(self,time):
		print("Total time taken is %s"%(time))
		t=[]
		x=s.goal
		while x!=-1:
			t.append(x)
			# print(x)
			x=self.p[x]
		for i in range(len(t)-1,-1,-1):
			print(t[i]+1,s.l[t[i]])

	def Distance(self,x,y,b):
		(X,Y)=(s.l[s.goal][0],s.l[s.goal][1])
		dist=math.sqrt( (x-X)**2 +(y-Y)**2)
		Tbus=10000.0
		if dist>3:
			Tbus=dist/s.Bus
			cost=Tbus*s.rate
			if b-cost<0:Tbus=10000.0
		Tcycl=dist/s.Cycle
		return min(Tbus,Tcycl)


	def BFS(self):
		heapq.heapify(self.h)
		(x,y)=(s.l[s.source][0],s.l[s.source][1])
		heapq.heappush(self.h,(0+self.Distance(x,y,s.budget),(s.source,-1,0,s.budget)))
		# heapq.heappush(self.h,(0,(s.source,-1,s.budget)))
		while self.h:
			(A,(node,P,g,b))=heapq.heappop(self.h)
			if self.v[node]==1:continue
			if self.v[node]==0:self.v[node]=1
			self.p[node]=P
			if node==s.goal:
				self.Print(g)
				break
			for i in range(0,s.n):
				n=s.d[node][i]
				if n!=-1.0:
					(x,y)=(s.l[i][0],s.l[i][1])
					Tbus=10000.0
					if n>3:
						Tbus=n/s.Bus
						cost=Tbus*s.rate
						if b-cost<0:Tbus=10000.0
						# heapq.heappush(self.h,(g+self.Distance(x,y,b),(i,node,g+t,b-cost)))
						# continue
					Tcycl=n/s.Cycle
					t=min(Tcycl,Tbus)
					if(t==Tcycl):heapq.heappush(self.h,(g+t+self.Distance(x,y,b),(i,node,g+t,b)))
					else: heapq.heappush(self.h,(g+t+self.Distance(x,y,b),(i,node,g+t,b-cost)))


s=Environment()
print("No congestion")
s.Bus=60
a=Agent()

print("\n50%s congestion"%("%"))
s.Bus=37.5
b=Agent()

print("\nFull congestion")
s.Bus=10
c=Agent()


